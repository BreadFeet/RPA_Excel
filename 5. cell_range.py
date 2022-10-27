from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# cell 한줄씩 데이터 넣기 - 성적표
ws.append(["번호", "영어", "수학"])   # 제목
for i in range(1, 11) : # 10개 번호 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

# 영어 점수만 추출 - column
col_B = ws["B"]    # B 컬럼만 가져옴
# print(col_B)     # 셀 정보 출력
# for cell in col_B :
#     print(cell.value)     # 셀 값 출력

# 여러 컬럼(영어, 수학) 추출
col_BC = ws["B:C"]
# for col in col_BC :
#     for cell in col :
#         print(cell.value)

# 번호 1만 추출 - row
row_2 = ws["2"]
# print(row_2)    # 정보만 추출
# for cell in row_2 :
#     print(cell.value)

# 제목 row만 가지고 오기
row_title = ws[1]     # row 숫자는, 숫자만 적어도 됨
# for title in row_title :
#     print(title.value)

# 여러 행 가져오기
row_26 = ws[2:6]     # 1~5번 학생 가져오기
# for row in row_26 : 
#     for cell in row :
#         print(cell.value, end=" ")
#     print()

# row, col 개수를 모를 때
row_range = ws[2:ws.max_row]   # 1번부터 마지막줄 학생까지
# for row in row_range :
#     for cell in row :
#         print(cell.value, end=" ")
#     print()

# cell의 value 대신 좌표 정보 가져오기
from openpyxl.utils.cell import coordinate_from_string

# for row in row_range :
#     for cell in row :
#         # print(cell.coordinate, end=" ")
#         xy = coordinate_from_string(cell.coordinate)    # ("A",10), ("AZ",250)...
#         # print(xy, end=" ")
#         print(xy[0], end=" ")   # column의 문자
#         print(xy[1], end=" ")   # row의 숫자
#     print()

# 전체 row 정보 가져오기
# print(tuple(ws.rows))

# for row in tuple(ws.rows) :  # tuple을 씌우지 않다고 결과는 같음
#     # print(row[1].value)    # 영어점수
#     print(row[0].value)      # 번호

# for row in ws.iter_rows() :    # 위와 같은 결과
#     print(row[1].value)      

# for row in ws.iter_rows(min_row=1, max_row=5) :   
#     print(row[1].value)    # 영어점수 1~4번

# 전체 columns 정보 가져오기
# print(tuple(ws.columns))

# for col in tuple(ws.columns) :
#     print(col[5].value)     # 5번 학생
# Excel cell은 1부터 시작, tuple은 0부터 시작!!

# for col in ws.iter_cols() :   # 위와 같은 결과
#       print(col[5].value)   

# 1~5번째줄, 2~3번째 열까지
# for cell in ws.iter_cols(min_row=1, max_row=5, min_col=2, max_col=3) :   
#     print(cell[2].value)    # 2번 학생의 영어, 수학점수

for cell in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3) :
    print(cell[0].value)    # 1~4번 영어점수

# 셀 정보를 바로 가져오기 - 편함
for cell in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3, values_only=True) :
    print(cell[0])    # 1~4번 영어점수

wb.save("sample.xlsx")