from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 영어점수 > 80 이상 학생
for row in ws.iter_rows(min_row=2) :
    if int(row[1].value) > 80 :   
        print(row[0].value, "번 학생은 영어를 잘하네요!")

# 영어점수가 아니라 컴퓨터 점수였다면,
for row in ws.iter_rows(max_row=1) :   # 첫번째 줄만
    for cell in row :
        if cell.value == "영어" :
            cell.value = "컴퓨터"

wb.save("sample_modified.xlsx")

