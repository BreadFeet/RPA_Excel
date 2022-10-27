from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = load_workbook("sample_formula.xlsx").active

# 엑셀 표 그대로 출력하기
# for row in ws.values :   # ws.values: 행으로 끊어서 출력
#     print(row)

# 수식으로 불러오기
# for row in ws.values :
#     for cell in row :
#         print(cell)

# 수식 계산된 데이터로 가져오기
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

for row in ws.values :
    for cell in row :
        print(cell)   
# None: 엑셀에 수식으로 쓰여있는 경우, 데이터는 엑셀이 계산해서 보여주는 거라
# openpyxl은 수식이지 데이터가 아니라고 생각해서 None
# evaluate 되지 않은 상태는 None이라고 표시
# 해결방법: 엑셀파일을 열어서 저장을 하면 evaluate된 값이 저장된다.

