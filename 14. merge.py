from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 셀 병합하기
ws.merge_cells("B2:D2")    # B2~D2 셀을 합침
ws["B2"] = "Merged Cell"   # C2, D2로 지정하면 오류남

wb.save("sample_merge.xlsx")