from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "MySheet"

# 셀에 값을 입력하기
ws["A1"] = 1     # A1 셀에 1 이라는 값을 입력
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

# 셀의 값을 읽기
print(ws["A1"])     # A1 셀의 "정보"를 출력
print(ws["A1"].value)     # A1 셀의 "값"을 출력
print(ws["A10"].value)   # 값이 없음: None 출력

print(ws.cell(row=1, column=1).value)  # 다른 방식으로 A1 출력
print(ws.cell(1, 2).value)        # B1 출력

c1 = ws.cell(1, 3, value=10)     # ws["C1"] = 10과 동일
print(c1.value)     # ws["C1"].value

from random import *

# interation 이용해서 랜덤 숫자 채우기
index = 1
for x in range(1, 11) :     # 10개 row
    for y in range(1, 11) :     # 10개 col 
        # ws.cell(x, y, value=randint(0, 100))  # 0~100 숫자 랜덤 지정
        
        # cell이 어느 순서로 채워지는지 확인
        ws.cell(x, y, value=index)
        index += 1        # 1행, 2행...순으로 채워짐


wb.save("sample.xlsx")   