from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 성적 데이터 입력하기
ws.append(["학번", "출석", '퀴즈1', '퀴즈2', '중간고사', '기말고사', '프로젝트'])
score = [[1,10,8,5,14,26,12],
[2,7,3,7,15,24,18],
[3,9,5,8,8,12,4],
[4,7,8,7,17,21,18],
[5,7,8,7,16,25,15],
[6,3,5,8,8,17,0],
[7,4,9,10,16,27,18],
[8,6,6,6,15,19,17],
[9,10,10,9,19,30,19],
[10,9,8,8,20,25,20]]

for x in score :
    ws.append(x)

# 퀴즈2 점수를 10으로 수정
# for cell in ws['D'] :
#     if isinstance(cell.value, int) :
#         cell.value = 10

# 나도코딩 방법
for idx, cell in enumerate(ws["D"]) :
    if idx == 0 :
        continue
    cell.value = 10


# H열에 총점 추가
ws["H1"] = "총점"     # 타이틀 넣기
ws["I1"] = "성적"

# # 총점 계산에 score 그대로 사용하면 안됨 - 퀴즈2가 10점으로 바뀌었으므로!
for idx, row in enumerate(score, start=2) :  # 엑셀표에서 2번째줄에 해당하므로 2에서 시작
    sum_val = sum(row[1:]) - row[3] + 10  # 퀴즈2=10점 기준 총점
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx) 
    # evaluate 되기 전까지 sum은 수식상태

    # 성적: 총점 90이상 A, 80이상 B, 70점 이상 C, 나머지 D
    if row[1] < 5 :
        ws.cell(row=idx, column=9).value = "F"
    elif sum_val >= 90 :
        ws.cell(row=idx, column=9).value = "A"
    elif 80 <= sum_val < 90 :
        ws.cell(row=idx, column=9).value = "B"
    elif 70 <= sum_val < 80 :
        ws.cell(row=idx, column=9).value = "C"
    else :
        ws.cell(row=idx, column=9).value = "D"
     

wb.save("scores.xlsx")
