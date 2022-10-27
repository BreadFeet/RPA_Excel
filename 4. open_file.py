from openpyxl import load_workbook    # 파일 불러오기
wb = load_workbook("sample.xlsx")    # sample 파일에서 wookbook을 불러옴
ws = wb.active    # 활서오하된 sheet를 불러온다

# cell 데이터 불러오기
# for x in range(1, 11) :    # 10개 row
#     for y in range(1, 11) :    # 10개 col
#         print(ws.cell(row=x, column=y).value, end=" ")  # 행 순서대로 일을테니까 1 2 3...
#     print()      # end는 무조건 " "로 계속 연결되는데 빈칸이 나오면 끊긴다 --> 다음줄 넘어감

# cell 갯수를 모를 때
for x in range(1, ws.max_row + 1) :  # +1을 해야 max_row까지 포함한다
    for y in range(1, ws.max_column + 1) :
        print(ws.cell(x, y).value, end="  ") 
    print()    # end는 무조건 " "로 계속 연결되는데 빈칸이 나오면 끊긴다.