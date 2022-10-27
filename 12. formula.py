import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today()  # 오늘 날짜정보
ws["A2"] = "=SUM(1,2,3)"    # 따옴표 안에 cell 입력 방식 그대로
ws["A3"] = "=AVERAGE(1,2,3)"

ws["A4"] = 10
ws["A5"] = 20
ws["A6"]= "=SUM(A4:A5)"     # 10+20

wb.save("sample_formula.xlsx")