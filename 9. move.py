from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 바로 옆에 "국어" 점수 삽입
# ws.move_range("B1:C11", rows=0, cols=1)   # 행은 그대로, 열만 한칸 옆으로 이동
# ws["B1"] .value = "국어"

# 수학을 영어 밑으로 옮기기
ws.move_range("C1:C11", rows=11, cols=-1)


wb.save("sample_kr.xlsx")