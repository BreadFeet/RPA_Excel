from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws["A1"]    # 번호
b1 = ws["B1"]    # 영어
c1 = ws["C1"]    # 수학

# A 컬럼의 폭을 줄이기
ws.column_dimensions["A"].width = 5

# 1행의 높이를 늘리기
ws.row_dimensions[1].height = 50

# 폰트 스타일 적용
from openpyxl.styles import Font
a1.font = Font(color="ff0150", italic=True, bold=True)
b1.font = Font(color="c3928f", name="Merriweather", strikethrough=True)
ws["C1"].font = Font(color="b9100f", size=20, underline="double") 

# 테두리 적용
from openpyxl.styles import Border, Side
# 테두리 스타일 설정
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thick"))
a1.border = thin_border    # 테두리 적용
b1.border = thin_border
ws["C1"].border = thin_border

# 90점이 넘는 경우 배경색 지정
from openpyxl.styles import PatternFill, Alignment

for row in ws.rows :
    for cell in row :
        # 셀 값 좌우 중앙, 상하 윗쪽 정렬
        cell.alignment = Alignment(horizontal="center", vertical="top")
        
        if cell.column == 1 :     # 번호 열은 필요없음
            continue              # skip
        
        # cell이 정수이고(제목제외), 90점 이상인 경우
        if isinstance(cell.value, int) and cell.value>90 :
            cell.fill = PatternFill(fgColor="00FF00", fill_type="lightGrid")
            cell.font = Font(color="FF0000")


# 제목 틀 고정(freeze)
ws.freeze_panes = "B2"     # B2 기준으로 윗줄, 왼쪽에 틀 고정


wb.save("sample_style.xlsx")