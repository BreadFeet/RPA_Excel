from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()     # 활성화된 sheet 뒤에 새로운 sheet 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor = "f155b0"   # RGB 형태로 탭색상 지정

# 새로운 시트 만들기
ws1 = wb.create_sheet("YourSheet")  # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet", 2)  # index=2 위치에 sheet 생성

# NewSheet 접근
# ws2...       # 파일 생성 변수를 통해 불러움
new_ws = wb["NewSheet"]   # Dictionary 형태로 sheet 불러옴
# new_ws...

print(wb.sheetnames)   # 모든 sheet 이름 출력

# Sheet 복사
new_ws["A1"] = "Text"  # A1 cell을 "Text"로 지정
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"


wb.save("sample.xlsx")